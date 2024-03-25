"""
Script designed for preparing excel request forms for branch managers to create order of products for their branches
"""

# import libraries
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill, Protection, Border, Side
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.datavalidation import DataValidation
import logging
import time
import traceback
from datetime import datetime as dt

# set up logging
logging.basicConfig(level=logging.DEBUG, encoding= 'utf-8',
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S',
                    filename='request_form.log',
                    filemode='w')


# file locations
EVALUATION_LOC = r'D:\Tasks\yoyoso restock planning\product evaluations\product_evaluation.csv' # location of product evaluation
SALES_LOC = r'D:\excel db\yoyoso\sales\sales_cleaned.csv' # location of sales data
INVENTORY_LOC = r'D:\excel db\yoyoso\inventory\inventory_clean.csv'
CLOSING_INVENTORY = r'D:\excel db\yoyoso\inventory\closing_inv_margins\closing_inventory_margins.xlsx' # grab codes from here
PRODUCT_DESCRIPTION = r'D:\excel db\yoyoso\product_description.xlsx'
REMOVE_CODES = r'D:\Tasks\yoyoso restock planning\restock branches\remove_codes\ბრუნვა.xlsx'

# directory of branch files
BRANCHES_DIR = r'D:\Tasks\yoyoso restock planning\restock branches\branches'

# get list of codes, that need to be removed
def remove_codes(code_dir: str) -> pd.DataFrame:
    code_list = pd.read_excel(code_dir, skiprows=1)
    code_list = code_list[['შიდა კოდი']]
    code_list.columns = ['code']
    code_list.code = code_list.code.astype('O')
    code_list.dropna(subset=['code'], inplace=True)
    
    return code_list

# read csv files and clean data
def prep_dataframes(evaluation_loc, sales_loc, inventory_loc, closing_inventory_loc, product_description_loc, centr_strg_name, warehouse_list):
    """
    product_evaluation, 
    sales_df, 
    inventory_df - closed inventory file, 
    closing_inventory, 
    central_storage_df, 
    share_of_sales_by_warehouses
    """
    
    # read csv files
    product_evaluation = pd.read_csv(evaluation_loc)
    sales_df = pd.read_csv(sales_loc)
    closing_inventory = pd.read_excel(closing_inventory_loc, skiprows=2)
    inventory_df = pd.read_csv(inventory_loc)
    product_description = pd.read_excel(product_description_loc)
    
    # clean product description
    product_description = product_description[['შიდა კოდი', 'რაოდენობა ყუთში']]
    product_description.rename({'შიდა კოდი': 'code', 'რაოდენობა ყუთში': 'box_quant'}, axis=1, inplace=True)
    
    # clean inventory_df
    inventory_df.sku = inventory_df.sku.astype('str')
    inventory_df['date'] = inventory_df.year.astype('str') + '-' + inventory_df.month.astype('str')
    inventory_df['date'] = pd.to_datetime(inventory_df['date'])
    
    # clean sales
    sales_df.sku = sales_df.sku.astype('str') # change `sku` type to str
    sales_df.date = pd.to_datetime(sales_df.date)
    
    # remove columns
    remove_columns = ['საწყობი', 'შიდა კოდი',
        'შტრიხკოდი', 'საქონელი', 'კატეგორია',
        'ტიპი', 'თვითღირებულება (Sum)', 'რაოდენობა (Sum)']
    closing_inventory = closing_inventory[remove_columns]
    
    # rename columns
    column_names = ['warehouse', 'code', 'sku', 'product_name', 'category', 'type', 'cogs', 'quantity']
    closing_inventory.columns = column_names
    
    # fill down warehouse
    closing_inventory.warehouse.ffill(inplace=True)
    
    # remove empty values
    closing_inventory = closing_inventory.loc[~closing_inventory.code.isna(), :].reset_index(drop=True)
    
    # change sku data type
    closing_inventory.sku = closing_inventory.sku.astype('str')
    
    # keep only necessary warehouses
    closing_inventory = closing_inventory[closing_inventory.warehouse.isin(warehouse_list)]
    
    # remove unnecessary categories
    remove_categories = ['საშობაო', 'დეკორაცია-აქსესუარები', 'საზაფხულო',
       'სათამაშოები', 'აქსესუარები', 'სახის მოვლა',
       'ტანის მოვლა', 'ჭურჭელი', 'სამზარეულო',
       'ჩანთები', 'ჰიგიენა', 'საკანცელარიო', 'აბაზანა', 'ტექსტილი',
       'საოჯახო აქსესუარები', 'მობილურის აქსესუარები',
       'კოსმეტიკის აქსესუარები', 'თმის აქსესუარები', 'ტექნიკა',
       'ფიტნესი ', 'ჰაერის არომათერაპია', 'ელემენტები',
       'ზამთრის აქსესუარები', 'თმის მოვლა', 'ბიჟუტერია ',
       'შინაური ცხოველები', 'გაზაფხული-შემოდგომის აქსესუარები', 'ჩუსტები',
       'სასაჩუქრე ნაკრები', 'კომპიუტერის აქსესუარები',
       'პარფიუმერია', 'კოსმეტიკა']

    closing_inventory = closing_inventory[closing_inventory.category.isin(remove_categories)].reset_index(drop=True)

    central_storage_df = closing_inventory[closing_inventory.warehouse == centr_strg_name].reset_index(drop=True)

    # add box quant
    closing_inventory = pd.merge(left=closing_inventory, right=product_description, on='code', how='left').reset_index(drop=False)

    # shares of sales by warehouses
    share_of_sales_by_warehouses = sales_df.groupby('warehouse').agg({
        'quantity': 'sum',
        'cogs': 'sum'
    }).reset_index(drop=False)
    share_of_sales_by_warehouses = share_of_sales_by_warehouses[share_of_sales_by_warehouses.warehouse != centr_strg_name]
    # calculate share of sales between warehouses
    total_cogs = share_of_sales_by_warehouses['cogs'].sum()

    share_of_sales_by_warehouses['share'] = round(share_of_sales_by_warehouses['cogs'] / total_cogs,2)
    share_of_sales_by_warehouses.drop(columns=['cogs', 'quantity'], axis=1, inplace=True)
    
    return product_evaluation, sales_df, inventory_df, closing_inventory, product_description, central_storage_df, share_of_sales_by_warehouses


# prepare form for each warehouse
def request_form(warehouse_var, closing_inventory, central_storage_name, product_evaluation, sales_df, share_of_sales_by_warehouses, central_storage_df, product_description_df):
    """
    warehouse_var, 
    closing_inventory, 
    central_storage_name, 
    product_evaluation, 
    sales_df, 
    share_of_sales_by_warehouses, 
    central_storage_df
    """
    # product description part of the code
    temp_df = closing_inventory.copy().loc[(closing_inventory.warehouse == central_storage_name) |
                      (closing_inventory.warehouse.isin(warehouse_var)), :]
    temp_df.drop(index=temp_df[temp_df.code.duplicated()].query('warehouse == @central_storage_name').index, inplace=True)

    temp_df.loc[temp_df.warehouse == central_storage_name, 'quantity'] = 0
    temp_df.loc[temp_df.warehouse == central_storage_name, 'cogs'] = 0
    # temp_df = temp_df.loc[:, ~temp_df.columns.isin(['cogs'])]
    
    # count of warehouses in the dataframe
    count_warehouses_in_dataframe = len(temp_df.warehouse.unique())
    
    # keep the untouched temp_df to use it later for cogs
    untouched_temp_df = temp_df.copy()
    
    # pivot to divide storage and shop inventory quantity
    temp_df = untouched_temp_df.pivot_table(index=['code', 'sku', 'product_name', 'category', 'type'], 
                                    columns='warehouse', 
                                    values='quantity', 
                                    aggfunc='sum', 
                                    fill_value=0).reset_index(drop=False)
    
    
    # drop central storage
    temp_df.drop(columns=central_storage_name, axis=1, inplace=True)

    """# rename to simplify names
    if count_warehouses_in_dataframe == 3:
        temp_df.rename(columns={
            warehouse_var[1]: 'საწყობი',
            warehouse_var[0]: 'მაღაზია'
        }, inplace=True)
    else:
        temp_df.rename(columns={
            warehouse_var[0]: 'მაღაზია'
        }, inplace=True)"""
    
    # merge with product evaluation
    temp_df = pd.merge(left=temp_df, right=product_evaluation[['code', 'DSI', 'ABC', 'doh', 'margin']], on='code', how='left').reset_index(drop=True)
    
    # get monthly average sales
    sales_by_warehouse = sales_df[sales_df.warehouse.isin(warehouse_var)]
    
    grouped_by_code_date = sales_by_warehouse.groupby(['date', 'code']).agg({'cogs': 'sum', 'quantity': 'sum'}).reset_index(drop=False)
    
    grouped_by_code_date['year'] = grouped_by_code_date.date.dt.year
    grouped_by_code_date['month'] = grouped_by_code_date.date.dt.month
    
    grouped_by_month_year = grouped_by_code_date.groupby(['year', 'month', 'code']).agg({'cogs': 'sum', 'quantity': 'sum'}).reset_index(drop=False)
    
    monthly_sales_by_products = grouped_by_month_year.groupby(['code']).agg({'cogs': 'mean', 'quantity': 'mean'}).reset_index(drop=False)
    
    monthly_sales_by_products.quantity = round(monthly_sales_by_products.quantity, 0)
    monthly_sales_by_products.cogs = round(monthly_sales_by_products.cogs, 2)
    
    # append average sales to closing_inventory
    temp_df = pd.merge(left=temp_df, right=monthly_sales_by_products, on='code', how='left', suffixes=('_current_warehouse', '_average sales (m)')).reset_index(drop=True)
    
    # calculate available quantity for branch from central storage
    for_temp_in_central_storage = central_storage_df[['code', 'quantity']]
    for_temp_in_central_storage.quantity = round(for_temp_in_central_storage.quantity * \
        share_of_sales_by_warehouses.loc[share_of_sales_by_warehouses.warehouse.isin(warehouse_var), 'share'].values[0],0)
        
    for_temp_in_central_storage.rename({'quantity': 'available'}, axis=1, inplace=True)
    temp_df = pd.merge(left=temp_df, right=for_temp_in_central_storage, on='code', how='left').reset_index(drop=True)
    temp_df.available.fillna(0, inplace=True)
    
    temp_df.rename({
    'code': "შიდა კოდი",
    'sku': "შტრიხკოდი",
    'product_name': "დასახელება",
    'category': "კატეგორია",
    'type': "ტიპი",
    'margin': "მარჟა",
    'quantity': "საშუალოდ ნავაჭრი",
    'available': "ხელმისაწვდომი"
        }, axis=1, inplace=True)
    
    temp_df.drop(columns='cogs', inplace=True)
    
    # combine storages' quantities into 1
    if len(warehouse_var) == 2:
        if count_warehouses_in_dataframe == 3:
            temp_df['მარაგი რაოდენობა'] = temp_df[warehouse_var[0]] + temp_df[warehouse_var[1]]
            temp_df.drop(columns=[warehouse_var[0], warehouse_var[1]], inplace=True)
        else:
            temp_df.rename(columns={warehouse_var[0]: 'მარაგი რაოდენობა'}, inplace=True)
    else:
        temp_df.rename(columns={warehouse_var[0]: 'მარაგი რაოდენობა'}, inplace=True)

    
    # pivot to divide storage and shop inventory cogs
    temp_df_cogs = untouched_temp_df.pivot_table(index=['code', 'sku', 'product_name', 'category', 'type'], 
                                    columns='warehouse', 
                                    values='cogs', 
                                    aggfunc='sum', 
                                    fill_value=0).reset_index(drop=False)

    # combine storages' cogs into 1
    if len(warehouse_var) == 2:
        if count_warehouses_in_dataframe == 3:
            temp_df_cogs['მარაგი თვითღირ.'] = temp_df_cogs[warehouse_var[0]] + temp_df_cogs[warehouse_var[1]]
            temp_df_cogs.drop(columns=[warehouse_var[0], warehouse_var[1]], inplace=True)
        else:
            temp_df_cogs.rename(columns={warehouse_var[0]: 'მარაგი თვითღირ.'}, inplace=True)
    else:
        temp_df_cogs.rename(columns={warehouse_var[0]: 'მარაგი თვითღირ.'}, inplace=True)

    temp_df = pd.merge(left=temp_df, right=temp_df_cogs[['code', 'მარაგი თვითღირ.']], left_on='შიდა კოდი', right_on='code', how='left').reset_index(drop=False)
    
    # merge box_quantity
    temp_df = pd.merge(left=temp_df, right=product_description_df, on='code', how='left')
    
    temp_df.rename(columns={'box_quant': 'ყუთში რაოდენობა'}, inplace=True)
    
    # add recommended quantity
    recommended_quantity = np.where(
        temp_df['ხელმისაწვდომი'] == 0, 
        0,  # If true, set recommended quantity to 0
        np.where(
            round((temp_df['საშუალოდ ნავაჭრი'] - temp_df['მარაგი რაოდენობა']) / temp_df['ყუთში რაოდენობა'], 0) * temp_df['ყუთში რაოდენობა'] >= temp_df['ხელმისაწვდომი'],
            round((temp_df['საშუალოდ ნავაჭრი'] - temp_df['მარაგი რაოდენობა']) / temp_df['ყუთში რაოდენობა'], 0) * temp_df['ყუთში რაოდენობა'],  # If true, use this calculation
            round(temp_df['ხელმისაწვდომი'] / temp_df['ყუთში რაოდენობა'], 0) * temp_df['ყუთში რაოდენობა']  # If false, use this calculation
        )
    )

    temp_df['რეკომენდირებული რაოდენობა'] = recommended_quantity

    # available products adjustment
    temp_df['ხელმისაწვდომი'] = np.where(temp_df['ხელმისაწვდომი'] < temp_df['ყუთში რაოდენობა'],0,
                                        round(temp_df['ხელმისაწვდომი'] / temp_df['ყუთში რაოდენობა'], 0) * temp_df['ყუთში რაოდენობა'])
    
    # set priorities
    good_dsi = 90
    good_doh = 180
    average_margin = 54.47
    
    """
    პრიორიტეტები - A, B, C, D
    """
    
    temp_df['პრიორიტეტულობა'] = np.where(
        (
            ((temp_df['ABC'] == 'A') & (temp_df['doh'] <= good_doh) & (temp_df['DSI'] <= good_dsi) & (temp_df['მარჟა'] >= average_margin)) |
            ((temp_df['ABC'] == 'B') & (temp_df['doh'] <= good_doh) & (temp_df['DSI'] <= good_dsi) & (temp_df['მარჟა'] >= average_margin))
        ),
        'A',  # Value if condition is true
        np.where(
            (
                ((temp_df['ABC'] == 'A') & (temp_df['DSI'] <= good_dsi)) |
                ((temp_df['ABC'] == 'B') & (temp_df['DSI'] <= good_dsi))
            ),
            'B',
            np.where(
                (
                    ((temp_df['ABC'] == 'C') & (temp_df['doh'] > good_doh) & (temp_df['DSI'] > good_dsi) & (temp_df['მარჟა'] < average_margin))
                ),
                'D',
                'C'
            )
        )
    )

    temp_df.to_excel('check_logic.xlsx', index=False)
    
    # reorder columns
    reorder_columns = ['შიდა კოდი',
    'შტრიხკოდი',
    'დასახელება',
    'კატეგორია',
    'ტიპი',
    'პრიორიტეტულობა',
    'საშუალოდ ნავაჭრი',
    'მარაგი თვითღირ.',
    'მარაგი რაოდენობა',
    'რეკომენდირებული რაოდენობა',
    'ყუთში რაოდენობა',
    'ხელმისაწვდომი']

    temp_df = temp_df[reorder_columns]
    
    # remove unnecessary codes
    rmv_codes_list = remove_codes(REMOVE_CODES)
    
    temp_df = temp_df[~temp_df['შიდა კოდი'].isin(rmv_codes_list.code)]
    
    return temp_df

# calculate last row of a table in excel
def calculate_last_row(dataframe):
    last_row = dataframe.shape[0]+1+20
    
    return last_row

# create excel file and format it
def initiate_excel_file():
    wb = Workbook()
    ws = wb.active
    
    return ws, wb

# format excel file
def format_excel_file(ws, last_row, warehouse):
    # table location
    table_range = f"C21:P{last_row}"
    # set table name
    table = Table(displayName='table', ref=table_range)
    ws.add_table(table)
    
    # change format of header cells
    header_cells = ws['C21':'P21']

    # change style to bold
    change_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    table_header_font_style = Font(bold=True)

    # apply the style changes to each cell in range
    for row in header_cells:
        for cell in row:
            cell.alignment = change_alignment
            cell.font = table_header_font_style
            
    # add validation for filling up
    validation = DataValidation(type="custom", formula1="=$O22<=$N22", showErrorMessage=True,
                                errorTitle="გადაჭარბებით მოთხოვნა",
                                error="მოთხოვნილი რაოდენობა ნაკლები ან ტოლი უნდა იყოს ხელმისაწვდომ რაოდენობაზე")
    ws.add_data_validation(validation)
    validation.add(f"N22:O{last_row}")
    
    # set title of for the excel file
    try:
        branch_name = warehouse[1].split(" - ")[1]
    except Exception as e:
        branch_name = warehouse[0].split(" - ")[1]


    title_font_style = Font(size=20, bold=True, color="00AE4F")

    ws['A1'].value = branch_name
    ws['A1'].font = title_font_style

    # column widths
    ws.column_dimensions['A'].width = 1
    ws.column_dimensions['B'].width = 1
    ws.column_dimensions['G'].width = 21
    ws.column_dimensions['C'].width = 21
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    ws.row_dimensions[2].height = 8
    ws.row_dimensions[9].height = 8
    # ws.row_dimensions[15].height = 8
    ws.row_dimensions[20].height = 8

    # change color of cells in range 'C21:P21' to '4F81BD' and font color to 'FFFFFF'

    cell_range = ws['C21:P21']

    # Define the fill color and font color
    fill_color = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    font_color = Font(color='FFFFFF')

    # Apply the fill and font color to the cells in the range
    for cell in cell_range:
        for c in cell:
            c.fill = fill_color
            c.font = font_color

    # freeze rows above 23
    # ws.freeze_panes = 'A22'
    """
    # format dsi number
    for row in range(22, last_row+1):
        cell = ws[f'H{row}']
        cell.number_format = "#,##0"
    """
    """
    # format margin numbers
    for row in range(22, last_row+1):
        cell = ws[f'K{row}']
        cell.number_format = "#,##0.00"
    """
    
    # set style for შევსება
    fill_color = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
    border_side = Side(style='thin', color='000000')  # Black color for borders
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for row in range(22, last_row + 1):
        cell = ws.cell(row=row, column=15)
        cell.fill = fill_color
        cell.border = cell_border

    # protect sheet
    # ws.protection = SheetProtection(autoFilter=True)

    ws.protection.sheet = True
    for row in range(21, last_row + 1):  # Adjust the range as needed
        ws.cell(row=row, column=15).protection = Protection(locked=False)

    prot = ws.protection
    prot.autoFilter = False

# fill in values
def populate_excel_file(ws, last_row, dataframe, inventory_df, warehouse):
    
    """
    ws - active sheet,
    last_row - calculate last_row,
    dataframe - final file,
    inventory_df - inventory with dates,
    warehouse - list of warehouses
    """
    start_row = 21
    start_column = 3

    # convert data to rows
    rows = dataframe_to_rows(dataframe, index=False, header=True)

    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, value in enumerate(row, start=start_column):
            ws.cell(row=r_idx, column=c_idx, value=value)

    fillin_name = 'შევსება'
    ws["O21"].value = fillin_name

    # add fill up column
    restock_name = 'განახლებული'
    ws["P21"].value = restock_name

    for row in range(22, last_row+1):
        ws[f"P{row}"] = f'=O{row} + K{row}'

    # set up informational section
    abc_section = 'ABC-ს გადანაწილება'
    quantity_to_add = 'დასამატებელი რაოდენობა'
    # terminology_explanation = 'ტერმინოლოგიის განმარტება'

    # Locations where you want to set the values
    cell_addresses = ['B3', 'B10']
    section_names = [abc_section, quantity_to_add]

    style_section_names = Font(size=12, bold=True)

    # Assign values and styles to specified cells
    for address, name in zip(cell_addresses, section_names):
        cell = ws[address]
        cell.value = name  # Set the value for each cell
        cell.font = style_section_names  # Apply the font style

    # set calculations for abc section
    subsection_headers = ['ABCD', 'რეკომენდაცია', 'არსებული', 'განახლებული']
    subsection_headers_address = ['C4', 'D4', 'E4', 'F4']
    subsection_headers_styles = Font(color='A6A6A6')

    for address, name in zip(subsection_headers_address, subsection_headers):
        cell = ws[address]
        cell.value = name
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = subsection_headers_styles

    # populate abc subsection
    abc_abc_names = ['A', 'B', 'C', 'D']
    abc_recommendations = [0.2, 0.5, 0.2, 0.1]

    # Assigning values to cells in column C (C6:C9) for abc_abc_names
    for i, value in enumerate(abc_abc_names, start=5):  # Starting from row 6
        ws[f'C{i}'].value = value

    percent_style = NamedStyle(name="percent_style", number_format="0%")
    # Assigning values to cells in column D (D6:D8) for abc_recommendations
    for i, value in enumerate(abc_recommendations, start=5):  # Starting from row 6
        cell = ws[f'D{i}']
        cell.value = value
        cell.style = percent_style

    # current abc allocation
    for row in range(5,9):
        cell = ws[f'E{row}']
        cell.value = f'=SUMIF(table[პრიორიტეტულობა],C{row},table[მარაგი რაოდენობა])/SUM(table[მარაგი რაოდენობა])'
        cell.style = percent_style

    # abc allocation after restocking
    for row in range(5,9):
        cell = ws[f'F{row}']
        cell.value = f'=SUMIF(table[პრიორიტეტულობა],C{row},table[განახლებული])/SUM(table[განახლებული])'
        cell.style = percent_style

    # recommendation about how many products to add
    max_cogs_in_wh_df = inventory_df[inventory_df.warehouse.isin(warehouse)].groupby('date')['cogs'].sum().reset_index(drop=False)
    max_cogs_wh = max_cogs_in_wh_df.cogs.max()

    ws["C11"].value = "მაქს ტევადობა"
    ws["C12"].value = "მინ რაოდენობა"
    ws["C13"].value = "განახლებული ნაშთი"
    ws["C14"].value = "მინ შესავსები"
    ws["C15"].value = "მაქს შესავსები"

    # set values
    ws["D11"].value = max_cogs_wh
    ws["D12"].value = 0.7 * max_cogs_wh
    ws["D13"].value = '=(SUM(table[მარაგი თვითღირ.]) / SUM(table[მარაგი რაოდენობა])) * SUM(table[განახლებული])'
    ws['D14'].value = '=D12 - D13'
    ws["D15"].value = '=D11 - D13'

    # change cell formats
    for row in range(11,16):
        cell = ws[f'D{row}']
        cell.number_format = "#,##0"
    """
    # terminology explanation
    ws["C17"].value = "DSI - რამდენ დღეში ამოიყიდება მარაგი"
    ws["C18"].value = "DOH - რამდენი დღეა მარაგში"
    ws["C19"].value = "A - მაღალი მოგება, B - საშუალო, C- დაბალი"
    """
# save excel file
def save_excel_file(wb, warehouse):
    file_name = warehouse[0].split(' - ')[1]
    wb.save(f'{BRANCHES_DIR}\{file_name}.xlsx')

def main():
    
    start_time = time.time()
    
    # central storage
    central_storage_name = '1610011100 - ცენტრალური საწყობი (ლილო)'

    # filter warehouses
    # დროებით ამოღებულია '1610000100 - პიქსელი საწყობი'
    warehouses_of_interest = [
    '1610000200 - მარჯანიშვილი საწყობი',
    '1610000500 - ბათუმი საწყობი',
    '1610010100 - პიქსელი - ფილიალი 1',
    '1610011100 - ცენტრალური საწყობი (ლილო)',
    '1610011400 - ისთ ფოინთი საწყობი',
    '1610020100 - მარჯანიშვილი - ფილიალი 2',
    '1610041100 - რუსთაველის - ფილიალი 8',
    '1610041500 - რუსთაველი 8 საწყობი',
    '1610050100 - ბათუმი მაღაზია',
    '1610070100 - თბილისი მოლი - ფილიალი 7',
    '1610071400 - თბილისი მოლი საწყობი',
    '1610080100 - ბათუმი XS - ფილიალი',
    '1610090100 - პეკინი',
    '1610990100 - პეკინი საწყობი',
    '1610100100 - ისთ ფოინთი - ფილიალი 10',
    '1610110100 - ყაზბეგი',
    '1610111400 - ყაზბეგი საწყობი']

    # pair warehouses
    # დროებით ამოღებულია '1610000100 - პიქსელი საწყობი', 
    warehouse_pairs = [
        ['1610010100 - პიქსელი - ფილიალი 1'],
        ['1610000200 - მარჯანიშვილი საწყობი', '1610020100 - მარჯანიშვილი - ფილიალი 2'],
        ['1610000500 - ბათუმი საწყობი', '1610080100 - ბათუმი XS - ფილიალი'],
        ['1610011400 - ისთ ფოინთი საწყობი', '1610100100 - ისთ ფოინთი - ფილიალი 10'],
        ['1610041500 - რუსთაველი 8 საწყობი', '1610041100 - რუსთაველის - ფილიალი 8'],
        ['1610050100 - ბათუმი მაღაზია'],
        ['1610071400 - თბილისი მოლი საწყობი', '1610070100 - თბილისი მოლი - ფილიალი 7'],
        ['1610111400 - ყაზბეგი საწყობი', '1610110100 - ყაზბეგი'],
        ['1610090100 - პეკინი', '1610990100 - პეკინი საწყობი']
        ]

    try:
        product_evaluation, sales_df, inventory_df, closing_inventory, product_description_df, central_storage_df, share_of_sales_by_warehouses = \
            prep_dataframes(EVALUATION_LOC, SALES_LOC, INVENTORY_LOC, CLOSING_INVENTORY, PRODUCT_DESCRIPTION, central_storage_name, warehouses_of_interest)
    except Exception as e:
        logging.warning(f'Problem with preparation of dataframes - {e}')

    for w in warehouse_pairs:
        logging.info(f'preparing warehouses: {w}')
        try:
            details = request_form(w, closing_inventory, central_storage_name, product_evaluation, sales_df, share_of_sales_by_warehouses, central_storage_df, product_description_df)
        except Exception as e:
            logging.warning(f'error in request form preperation: {e}')
            logging.warning(f'failed: {w} - {traceback.format_exc()}')
            continue
        
        last_row = calculate_last_row(details)
        
        try:
            ws, wb = initiate_excel_file()
        except Exception as e:
            logging.warning((f'Problem with initiating excel file - {e}'))
            logging.warning(f'failed: {w} - {traceback.format_exc()}')
            continue
        
        # details = details.loc[:, ~details.columns.isin(['ყუთში რაოდენობა'])]
        
        try:
            populate_excel_file(ws, last_row, details, inventory_df, w)
        except Exception as e:
            logging.warning(f'Problem with population of excel file - {e}')
            logging.warning(f'failed: {w}')
            continue
        
        try:
            format_excel_file(ws, last_row, w)
        except Exception as e:
            logging.warning(f'Problem with formating of excel file - {e} - {traceback.format_exc()}')
            logging.warning(f'failed: {w}')
            continue
        
        try:
            save_excel_file(wb, w)
        except Exception as e:
            logging.warning(f'Problem with saving of excel file - {e}')
            logging.warning(f'failed: {w}')
            continue
        
        logging.info(f'{w} - prepared')
        
        end_time = time.time()
        passed_time = end_time - start_time
        logging.info(f"Execution time: {passed_time:.2f} seconds")

if __name__ == '__main__':
    main()
